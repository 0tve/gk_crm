import datetime
import logging
from typing import Any, Dict

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LoginView, PasswordResetConfirmView
from django.db import models
from django.db.models import Prefetch
from django.db.models.query import QuerySet
from django.forms import modelformset_factory, inlineformset_factory
from django.forms.models import BaseModelForm
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect
from django.shortcuts import redirect, render
from django.urls import reverse
from django.views import generic
from openpyxl import Workbook
from openpyxl.styles import Font
from csv import DictReader
import pandas


from agents.mixins import OrganisorAndLoginRequiredMixin

from .forms import (AssignAgentForm, CategoryModelForm, CustomLoginForm,
                    CustomPasswordResetConfirmForm, CustomUserCreationForm,
                    FollowUpModelForm, LeadModelForm,
                    OrderModelForm, LeadImport,
                    )
from .models import Category, FollowUp, Lead, Order, Agent

logger = logging.getLogger(__name__)


def export_leads_to_excel(request):
    # Получаем все объекты модели Lead с предзагрузкой связанных моделей Order и Category
    leads = Lead.objects.prefetch_related(
        Prefetch('order_set', queryset=Order.objects.select_related('category'))
    ).all()

    # Создаем новую книгу Excel
    workbook = Workbook()
    worksheet = workbook.active

    # Добавляем заголовки столбцов
    headers = ['Фамилия', 'Имя', 'Отчество', 'Возраст', 'Менеджер', 'Ответственный','Адрес', 'Телефон', 'Почта', 'В архиве', 'Заказ', 'Категория']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header

    # Добавляем данные о каждом лице и его заказах в таблицу
    row_num = 2
    for lead in leads:
        orders = lead.order_set.all()
        if orders:
            for index, order in enumerate(orders):
                row_data = []
                if index == 0:
                    row_data.extend([
                        lead.last_name,
                        lead.first_name,
                        lead.middle_name,
                        lead.age,
                        str(lead.organisation),
                        str(lead.agent),
                        lead.address,
                        lead.phone_number,
                        lead.email,
                        lead.is_active,
                    ])
                else:
                    row_data.extend([''] * 10)  # Add empty values for lead's information

                row_data.extend([
                    order.name,
                    str(order.category),
                ])

                for col_num, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    # Изменяем отображение логических значений
                    if isinstance(value, bool):
                        if value:
                            cell.value = "Нет"  # Или любое другое отображение для True
                        else:
                            cell.value = "Да"  # Или любое другое отображение для False
                    else:
                        cell.value = value
                row_num += 1
        else:
            row_data = [
                lead.last_name,
                lead.first_name,
                lead.middle_name,
                lead.age,
                str(lead.organisation),
                str(lead.agent),
                lead.address,
                lead.phone_number,
                lead.email,
                lead.is_active,
            ] + ['', '']  # Add empty values for order and category

            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                # Изменяем отображение логических значений
                if isinstance(value, bool):
                    if value:
                        cell.value = "Нет"  # Или любое другое отображение для True
                    else:
                        cell.value = "Да"  # Или любое другое отображение для False
                else:
                    cell.value = value
            row_num += 1

    # Устанавливаем ширину столбцов для лучшей читаемости
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Получаем буквенное обозначение столбца
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Множитель 1.2 для учета ширины символов
        worksheet.column_dimensions[column].width = adjusted_width

    # Добавляем фильтры к заголовкам столбцов
    worksheet.auto_filter.ref = worksheet.dimensions

    # Устанавливаем стиль для заголовков
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = header_font

    # Создаем HTTP-ответ с прикрепленным файлом Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=leads_report.xlsx'
    workbook.save(response)

    return response


def import_leads(file_path, manager):
    # Чтение данных из Excel-файла
    df = pandas.read_excel(file_path)

    for _, row in df.iterrows():
        # Получение значений из соответствующих столбцов
        first_name = row['Имя']
        middle_name = row['Отчество']
        last_name = row['Фамилия']
        age = row['Возраст']
        agent_name = row['Ответственный']
        manager = row['Менеджер']
        address = row['Адрес']
        phone_number = str(row['Телефон'])
        email = row['Почта']

        # Поиск или создание связанного объекта "Агент"
        agent, _ = Agent.objects.get_or_create(id=7)

        # Создание объекта Lead
        lead = Lead(
            first_name=first_name,
            middle_name=middle_name,
            last_name=last_name,
            age=age,
            organisation_id=11,  # Здесь нужно указать связанный объект "UserProfile"
            address=address,
            phone_number='+'+phone_number,
            email=email,
            agent=agent
        )

        # Сохранение объекта Lead
        try:
            lead.save()
        except Exception as e:
            # Обработка ошибок сохранения
            print(f"Ошибка сохранения: {e}")

    print("Импорт данных завершен.")


def import_leads_from_excel(request):
    if request.method == 'POST':
        form = LeadImport(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            manager = request.user.userprofile
            import_leads(excel_file, manager)
            return render(request, 'leads/lead_list.html')
    else:
        form = LeadImport()
    
    return render(request, 'leads/lead_import.html', {'form': form})


class OrderUpdateView(OrganisorAndLoginRequiredMixin, generic.UpdateView):
    template_name = 'leads/lead_orders.html'
    form_class = OrderModelForm

    def get_queryset(self):
        return Lead.objects.filter(id=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        OrderFormSet = inlineformset_factory(Lead, Order, form=OrderModelForm, extra=1)
        lead = Lead.objects.get(id=self.kwargs['pk'])
        formset = OrderFormSet(instance=lead)
        if self.request.method == "POST":
            formset = OrderFormSet(self.request.POST, instance=lead)
            if formset.is_valid():
                formset.save()

        context = {'formset': formset,
                   'lead': lead}
        return context

    def get_success_url(self):
        return reverse("leads:lead-detail", kwargs={"pk": self.get_object().id})


def lead_toggle_active(request, pk):
    lead = Lead.objects.get(id=pk)
    if lead.is_active:
        lead.is_active = False
    else:
        lead.is_active = True
    lead.save()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = "registration/password_reset_confirm.html"
    form_class = CustomPasswordResetConfirmForm


class CustomLoginView(LoginView):
    template_name = 'registration/login.html'
    form_class = CustomLoginForm


class SignupView(generic.CreateView):
    template_name = "registration/signup.html"
    form_class = CustomUserCreationForm

    def get_success_url(self):
        return reverse("login")


class LandingPageView(generic.TemplateView):
    template_name = "landing.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect("leads:lead-list")
        return super().dispatch(request, *args, **kwargs)


# class DashboardView(OrganisorAndLoginRequiredMixin, generic.TemplateView):
#     template_name = "dashboard.html"

#     def get_context_data(self, **kwargs):
#         context = super(DashboardView, self).get_context_data(**kwargs)

#         user = self.request.user

#         total_lead_count = Lead.objects.filter(organisation=user.userprofile).count()

#         thirty_days_ago = datetime.date.today() - datetime.timedelta(days=30)

#         total_in_past30 = Lead.objects.filter(
#             organisation=user.userprofile,
#             date_added__gte=thirty_days_ago
#         ).count()

#         try:
#             converted_category = Category.objects.get(name="Converted")
#         except Category.DoesNotExist:
#             converted_category = None
#         converted_in_past30 = Lead.objects.filter(
#             organisation=user.userprofile,
#             category=converted_category,
#             converted_date__gte=thirty_days_ago
#         ).count()

#         context.update({
#             "total_lead_count": total_lead_count,
#             "total_in_past30": total_in_past30,
#             "converted_in_past30": converted_in_past30
#         })
#         return context


class LeadListView(LoginRequiredMixin, generic.ListView):
    template_name = "leads/lead_list.html"
    context_object_name = "leads"

    def get_queryset(self):
        user = self.request.user
        if user.is_organisor:
            queryset = Lead.objects.filter(
                organisation=user.userprofile,
                agent__isnull=False,
                is_active=True
            )
        else:
            queryset = Lead.objects.filter(
                organisation=user.agent.organisation,
                agent__isnull=False,
                is_active=True
            )
            queryset = queryset.filter(agent__user=user)
        return queryset

    def get_context_data(self, **kwargs):
        context = super(LeadListView, self).get_context_data(**kwargs)
        user = self.request.user
        if user.is_organisor:
            queryset = Lead.objects.filter(
                organisation=user.userprofile,
                agent__isnull=True,
                is_active=True
            )
            context.update({
                "unassigned_leads": queryset
            })
        return context


class LeadArchiveView(LoginRequiredMixin, generic.ListView):
    template_name = "leads/lead_archive.html"
    context_object_name = "leads"

    def get_queryset(self):
        user = self.request.user
        if user.is_organisor:
            queryset = Lead.objects.filter(
                organisation=user.userprofile,
                agent__isnull=False,
                is_active=False
            )
        else:
            queryset = Lead.objects.filter(
                organisation=user.agent.organisation,
                agent__isnull=False,
                is_active=False
            )
            queryset = queryset.filter(agent__user=user)
        return queryset


class UnassignedLeadListView(LoginRequiredMixin, generic.ListView):
    template_name = "leads/unassigned_lead_list.html"
    context_object_name = "unassigned_leads"

    def get_queryset(self):
        user = self.request.user
        if user.is_organisor:
            queryset = Lead.objects.filter(
                organisation=user.userprofile,
                agent__isnull=True,
                is_active=True,
            )
        return queryset


class LeadDetailView(LoginRequiredMixin, generic.DetailView):
    template_name = "leads/lead_detail.html"
    context_object_name = "lead"

    def get_queryset(self):
        user = self.request.user
        if user.is_organisor:
            queryset = Lead.objects.filter(organisation=user.userprofile,
                                           is_active=True)
        else:
            queryset = Lead.objects.filter(
                organisation=user.agent.organisation,
                is_active=True)
            queryset = queryset.filter(agent__user=user)
        return queryset


class LeadCreateView(OrganisorAndLoginRequiredMixin, generic.CreateView):
    template_name = "leads/lead_create.html"
    form_class = LeadModelForm

    def get_form_kwargs(self):
        kwargs = super(LeadCreateView, self).get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_success_url(self):
        return reverse("leads:lead-list")

    def form_valid(self, form):
        user = self.request.user
        lead = form.save(commit=False)
        lead.organisation = self.request.user.userprofile

        if not Category.objects.filter(organisation=user.userprofile):
            Category.objects.create(name="Заказ оформлен", organisation=user.userprofile)

        lead.category = Category.objects.get(name='Заказ оформлен')
        lead.save()
        messages.success(self.request, 'Контакт был успешно создан')
        return super(LeadCreateView, self).form_valid(form)


class LeadUpdateView(OrganisorAndLoginRequiredMixin, generic.UpdateView):
    template_name = "leads/lead_update.html"
    form_class = LeadModelForm

    def get_form_kwargs(self):
        kwargs = super(LeadUpdateView, self).get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_queryset(self):
        user = self.request.user
        return Lead.objects.filter(organisation=user.userprofile, is_active=True)

    def get_success_url(self):
        return reverse("leads:lead-detail", kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        form.save()
        messages.info(self.request, "Контакт был успешно изменен")
        return super(LeadUpdateView, self).form_valid(form)


class LeadDeleteView(OrganisorAndLoginRequiredMixin, generic.DeleteView):
    template_name = "leads/lead_delete.html"

    def get_success_url(self):
        return reverse("leads:lead-list")

    def get_queryset(self):
        user = self.request.user
        return Lead.objects.filter(organisation=user.userprofile, is_active=True)


class AssignAgentView(OrganisorAndLoginRequiredMixin, generic.FormView):
    template_name = "leads/assign_agent.html"
    form_class = AssignAgentForm

    def get_form_kwargs(self, **kwargs):
        kwargs = super(AssignAgentView, self).get_form_kwargs(**kwargs)
        kwargs.update({
            "request": self.request
        })
        return kwargs

    def get_success_url(self):
        return reverse("leads:unassigned-lead-list")

    def form_valid(self, form):
        agent = form.cleaned_data["agent"]
        lead = Lead.objects.get(id=self.kwargs["pk"])
        lead.agent = agent
        lead.save()
        return super(AssignAgentView, self).form_valid(form)


class CategoryListView(LoginRequiredMixin, generic.ListView):
    template_name = "leads/category_list.html"
    context_object_name = "categories"

    def get_queryset(self):
        user = self.request.user

        if user.is_organisor:
            queryset = Category.objects.filter(
                organisation=user.userprofile
            )
        else:
            queryset = Category.objects.filter(
                organisation=user.agent.organisation
            )
        return queryset 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        orders_by_category = {}
        queryset = self.get_queryset()
        for category in queryset:
            orders = Order.objects.filter(category=category)
            orders_by_category[category] = orders
            
        context['orders_by_category'] = orders_by_category
        return context


class CategoryDetailView(LoginRequiredMixin, generic.DetailView):
    template_name = "leads/category_detail.html"
    context_object_name = "category"

    def get_queryset(self):
        return Category.objects.filter(id=self.kwargs['pk'])
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        category = Category.objects.get(id=self.kwargs['pk'])
        orders = Order.objects.filter(category=category)
        
        context.update({'orders': orders})
        return context
    
    
# class CategoryDetailView(LoginRequiredMixin, generic.DetailView):
#     template_name = "leads/category_detail.html"
#     context_object_name = "category"

#     def get_queryset(self):
#         user = self.request.user
#         if user.is_organisor:
#             queryset = Category.objects.filter(
#                 organisation=user.userprofile
#             )
#         else:
#             queryset = Category.objects.filter(
#                 organisation=user.agent.organisation
#             )
#         return queryset


class CategoryCreateView(OrganisorAndLoginRequiredMixin, generic.CreateView):
    template_name = "leads/category_create.html"
    form_class = CategoryModelForm

    def get_success_url(self):
        return reverse("leads:category-list")

    def form_valid(self, form):
        category = form.save(commit=False)
        category.organisation = self.request.user.userprofile
        category.save()
        return super(CategoryCreateView, self).form_valid(form)


class CategoryUpdateView(OrganisorAndLoginRequiredMixin, generic.UpdateView):
    template_name = "leads/category_update.html"
    form_class = CategoryModelForm

    def get_success_url(self):
        return reverse("leads:category-detail", kwargs={'pk': self.object.pk})

    def get_queryset(self):
        user = self.request.user
        if user.is_organisor:
            queryset = Category.objects.filter(
                organisation=user.userprofile
            )
        else:
            queryset = Category.objects.filter(
                organisation=user.agent.organisation
            )
        return queryset


class CategoryDeleteView(OrganisorAndLoginRequiredMixin, generic.DeleteView):
    template_name = "leads/category_delete.html"

    def get_success_url(self):
        return reverse("leads:category-list")

    def get_queryset(self):
        user = self.request.user
        if user.is_organisor:
            queryset = Category.objects.filter(
                organisation=user.userprofile
            )
        else:
            queryset = Category.objects.filter(
                organisation=user.agent.organisation
            )
        return queryset


# class LeadCategoryUpdateView(LoginRequiredMixin, generic.UpdateView):
#     template_name = "leads/lead_category_update.html"
#     form_class = LeadCategoryUpdateForm

#     def get_queryset(self):
#         user = self.request.user
#         if user.is_organisor:
#             queryset = Lead.objects.filter(organisation=user.userprofile, is_active=True)
#         else:
#             queryset = Lead.objects.filter(
#                 organisation=user.agent.organisation,
#                 is_active=True)
#             queryset = queryset.filter(agent__user=user)
#         return queryset

#     def get_success_url(self):
#         return reverse("leads:lead-detail", kwargs={"pk": self.get_object().id})

#     def form_valid(self, form):
#         lead_before_update = self.get_object()
#         instance = form.save(commit=False)
#         try:
#             converted_category = Category.objects.get(name='Договор заключен')
#         except:
#             converted_category = None
#         if form.cleaned_data['category'] == converted_category:
#             if lead_before_update.category != converted_category:
#                 instance.converted_date = datetime.datetime.now()
#         instance.save()
#         messages.info(self.request, 'Статус был успешно изменен')
#         return super(LeadCategoryUpdateView, self).form_valid(form)


class FollowUpCreateView(LoginRequiredMixin, generic.CreateView):
    template_name = "leads/followup_create.html"
    form_class = FollowUpModelForm

    def get_success_url(self):
        return reverse("leads:lead-detail", kwargs={'pk': self.kwargs['pk']})

    def get_context_data(self, **kwargs):
        context = super(FollowUpCreateView, self).get_context_data(**kwargs)
        context.update({
            'lead': Lead.objects.get(pk=self.kwargs['pk'])
        })
        return context

    def form_valid(self, form):
        lead = Lead.objects.get(pk=self.kwargs['pk'])
        followup = form.save(commit=False)
        followup.lead = lead
        followup.save()
        return super(FollowUpCreateView, self).form_valid(form)


class FollowUpUpdateView(LoginRequiredMixin, generic.UpdateView):
    template_name = "leads/followup_update.html"
    form_class = FollowUpModelForm

    def get_queryset(self):
        user = self.request.user
        if user.is_organisor:
            queryset = FollowUp.objects.filter(
                lead__organisation=user.userprofile)
        else:
            queryset = FollowUp.objects.filter(
                lead__organisation=user.agent.organisation)
            queryset = queryset.filter(lead__agent__user=user)
        return queryset

    def get_success_url(self):
        return reverse("leads:lead-detail", kwargs={"pk": self.get_object().lead.id})


class FollowUpDeleteView(OrganisorAndLoginRequiredMixin, generic.DeleteView):
    template_name = "leads/followup_delete.html"

    def get_success_url(self):
        followup = FollowUp.objects.get(id=self.kwargs['pk'])
        return reverse("leads:lead-detail", kwargs={'pk': followup.lead.pk})

    def get_queryset(self):
        user = self.request.user
        if user.is_organisor:
            queryset = FollowUp.objects.filter(
                lead__organisation=user.userprofile)
        else:
            queryset = FollowUp.objects.filter(
                lead__organisation=user.agent.organisation)
            queryset = queryset.filter(lead__agent__user=user)
        return queryset
